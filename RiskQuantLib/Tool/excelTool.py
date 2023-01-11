#!/usr/bin/python
#coding = utf-8

import pandas as pd
import openpyxl
#<import>
#</import>

def formatExcelWithTemplate(templateExcelPathString:str,dataExcelPathString:str,savePathString:str):
    """
    This function will use the format of templateExcel and the data of dataExcel to
    generate a new excel.

    Parameters
    ----------
    templateExcelPathString : str
        The xlsx file that specify the format of output.
    dataExcelPathString : str
        The xlsx file that specify the content of output.
    savePathString : str
        The path where the output should be saved.

    Returns
    -------
    None
    """
    columnNameList = [chr(i).upper() for i in range(97,123)]

    # import template
    wb = openpyxl.load_workbook(templateExcelPathString)
    sheetnames = wb.get_sheet_names()


    # load data
    wbOfData = openpyxl.load_workbook(dataExcelPathString)
    sheetnamesOfData = wbOfData.get_sheet_names()

    for tmp,data in zip(sheetnames,sheetnamesOfData):
        sheet = wb.get_sheet_by_name(tmp)
        sheetOfData = wbOfData.get_sheet_by_name(data)

        # replace content
        max_column = max([sheet.max_column,sheetOfData.max_column])
        max_row = max([sheet.max_row,sheetOfData.max_row])
        for col in range(max_column):
            for row in range(max_row):
                blockIndexString = columnNameList[col]+str(row+1)
                sheet[blockIndexString] = sheetOfData[blockIndexString].value
    wb.save(savePathString)

#<excelTool>
#</excelTool>






